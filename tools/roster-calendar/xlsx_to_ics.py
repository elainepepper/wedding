#!/usr/bin/env python3
"""Convert the Tech Roster workbook into an iCalendar (.ics) file for one person.

This is the offline twin of roster-ics.gs: same block-detection and same
colour -> shift mapping, but reading an .xlsx export instead of the live
Google Sheet. Use it for a one-off import or to sanity-check the mapping;
use the Apps Script for the auto-updating subscription.

    python3 xlsx_to_ics.py Tech_Roster.xlsx --person Elaine --out elaine.ics

The roster stores the shift *time* in the cell's fill colour (see the legend
at AJ3:AJ9 of a year tab); cell text is hours/notes, not the shift.
"""

import argparse
import datetime as dt
import re
import sys

import openpyxl

TIMEZONE = "Australia/Perth"

MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
MONTH_NUM = {m.lower(): i for i, m in enumerate(MONTHS, 1)}
WEEKDAYS = {"mon", "tue", "wed", "thu", "fri", "sat", "sun"}

# Fill colour -> shift. Times are (start_h, start_m, end_h, end_m); an end that
# is not after the start rolls over to the next day (night shift).
SHIFTS = {
    "95b3d7": ("8am - 4pm", (8, 0), (16, 0)),
    "ffd966": ("8.30am - 4.30pm", (8, 30), (16, 30)),
    "90dcbf": ("8am - 5pm", (8, 0), (17, 0)),
    "00b050": ("9am - 5pm", (9, 0), (17, 0)),
    "ff9900": ("2pm - 10pm", (14, 0), (22, 0)),
    "3366ff": ("4pm - midnight", (16, 0), (0, 0)),
    "7030a0": ("Night 9.30pm - 8am", (21, 30), (8, 0)),
    "c65911": ("1pm - 9pm", (13, 0), (21, 0)),
}

# Fill colour -> all-day entry (not a worked shift).
NON_SHIFTS = {
    "434343": "Annual leave",
    "cccccc": "RDO / day in lieu",
}

# Cell text -> all-day entry, used when the cell has no meaningful fill.
TEXT_ENTRIES = {
    "phol": "Public holiday",
    "phoe": "Public holiday",
    "a/l": "Annual leave",
    "al": "Annual leave",
}

WHITE = {"", "ffffff", "00000000", "none"}

# "8-5", "2-10", "9.30-2", "12.30-9.30", "10-6" ...
TIME_RANGE = re.compile(
    r"^\s*(\d{1,2})(?:[.:](\d{2}))?\s*(am|pm)?\s*-\s*(\d{1,2})(?:[.:](\d{2}))?\s*(am|pm)?\s*$",
    re.I,
)


def norm_colour(cell):
    """Return the cell's fill as a bare lowercase rrggbb, or '' when unfilled."""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return ""
    rgb = getattr(fill.start_color, "rgb", None)
    if not isinstance(rgb, str):
        return ""
    rgb = rgb.lower()
    if len(rgb) == 8:  # aarrggbb
        rgb = rgb[2:]
    return "" if rgb in WHITE else rgb


def to_24h(hour, minute, meridiem, *, is_end, start_minutes=None):
    """Best-effort am/pm inference for hand-typed ranges like '8-5' or '2-10'."""
    if meridiem:
        hour = hour % 12 + (12 if meridiem.lower() == "pm" else 0)
    elif not is_end:
        # Shifts start between 06:00 and 11:59, otherwise it's an afternoon start.
        if hour < 6:
            hour += 12
    else:
        minutes = hour * 60 + minute
        if start_minutes is not None and minutes <= start_minutes and hour < 12:
            hour += 12
    return hour % 24, minute


def parse_time_range(text):
    m = TIME_RANGE.match(text)
    if not m:
        return None
    sh, sm, sap, eh, em, eap = m.groups()
    sh, sm = to_24h(int(sh), int(sm or 0), sap, is_end=False)
    eh, em = to_24h(int(eh), int(em or 0), eap, is_end=True, start_minutes=sh * 60 + sm)
    return (sh, sm), (eh, em)


def cell_text(value):
    if value is None or isinstance(value, (dt.datetime, dt.date)):
        return ""
    if isinstance(value, float):
        return ("%g" % value)
    return str(value).strip()


def find_blocks(ws, max_col):
    """Locate each month grid: (day_number_row, weekday_row, month, year_offset)."""
    blocks = []
    prev_month = 0
    year_offset = 0
    prev_row = 0
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 2).value != 1 or ws.cell(r, 3).value != 2:
            continue
        below = ws.cell(r + 1, 2).value
        if not isinstance(below, str) or below.strip().lower()[:3] not in WEEKDAYS:
            continue
        month = 0
        for rr in range(max(prev_row + 1, r - 4), r):
            for cc in range(1, max_col + 1):
                v = ws.cell(rr, cc).value
                if isinstance(v, str) and v.strip().lower() in MONTH_NUM:
                    month = MONTH_NUM[v.strip().lower()]
        if not month:
            month = prev_month + 1 if prev_month else 1
            if month > 12:
                month = 1
        if month < prev_month:  # rolled into the following year
            year_offset += 1
        blocks.append((r, r + 1, month, year_offset))
        prev_month, prev_row = month, r
    return blocks


def person_row(ws, block_row, next_block_row, person):
    limit = next_block_row if next_block_row else ws.max_row + 1
    for r in range(block_row + 2, min(limit, ws.max_row + 1)):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.strip().lower() == person.lower():
            return r
    return None


def collect(ws, year, person, max_col=32):
    """Yield one record per rostered day for `person` on this year tab."""
    blocks = find_blocks(ws, max_col)
    for i, (day_row, _wd_row, month, offset) in enumerate(blocks):
        nxt = blocks[i + 1][0] if i + 1 < len(blocks) else None
        prow = person_row(ws, day_row, nxt, person)
        if prow is None:
            continue
        for c in range(2, max_col + 1):
            day = ws.cell(day_row, c).value
            if not isinstance(day, (int, float)):
                continue
            cell = ws.cell(prow, c)
            colour = norm_colour(cell)
            text = cell_text(cell.value)
            if not colour and text.lower() in ("", "dod"):
                continue
            try:
                date = dt.date(year + offset, month, int(day))
            except ValueError:
                continue
            yield date, colour, text


def build_events(records, include_leave=True):
    events = []
    for date, colour, text in records:
        note = text if not re.fullmatch(r"-?\d+(\.\d+)?", text) else ""
        explicit = parse_time_range(text) if text else None

        if colour in SHIFTS or explicit:
            if explicit:
                label = text
                (sh, sm), (eh, em) = explicit
            else:
                label, (sh, sm), (eh, em) = SHIFTS[colour]
            start = dt.datetime(date.year, date.month, date.day, sh, sm)
            end = dt.datetime(date.year, date.month, date.day, eh, em)
            if end <= start:
                end += dt.timedelta(days=1)
            events.append({
                "kind": "timed",
                "date": date,
                "start": start,
                "end": end,
                "summary": "Work: " + label,
                "note": note if note != label else "",
            })
        elif colour in NON_SHIFTS:
            if include_leave:
                events.append({
                    "kind": "allday", "date": date,
                    "summary": NON_SHIFTS[colour], "note": note,
                })
        elif text.lower() in TEXT_ENTRIES:
            if include_leave:
                events.append({
                    "kind": "allday", "date": date,
                    "summary": TEXT_ENTRIES[text.lower()], "note": note,
                })
        elif colour:
            events.append({
                "kind": "allday", "date": date,
                "summary": "Rostered - check the sheet",
                "note": ("%s (colour #%s)" % (note, colour)).strip(),
            })
    return events


def esc(text):
    return (text.replace("\\", "\\\\").replace(";", r"\;")
                .replace(",", r"\,").replace("\n", r"\n"))


def fold(line):
    out, raw = [], line.encode("utf-8")
    while len(raw) > 73:
        out.append(raw[:73].decode("utf-8", "ignore"))
        raw = raw[73:]
    out.append(raw.decode("utf-8", "ignore"))
    return "\r\n ".join(out)


def to_ics(events, person, stamp):
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//roster-calendar//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:%s - Work Roster" % person,
        "X-WR-TIMEZONE:" + TIMEZONE,
        "REFRESH-INTERVAL;VALUE=DURATION:PT4H",
        "X-PUBLISHED-TTL:PT4H",
        "BEGIN:VTIMEZONE",
        "TZID:" + TIMEZONE,
        "BEGIN:STANDARD",
        "DTSTART:19700101T000000",
        "TZOFFSETFROM:+0800",
        "TZOFFSETTO:+0800",
        "TZNAME:AWST",
        "END:STANDARD",
        "END:VTIMEZONE",
    ]
    for ev in events:
        uid = "%s-%s@roster" % (ev["date"].strftime("%Y%m%d"), person.lower())
        lines += ["BEGIN:VEVENT", "UID:" + uid, "DTSTAMP:" + stamp]
        if ev["kind"] == "timed":
            lines += [
                "DTSTART;TZID=%s:%s" % (TIMEZONE, ev["start"].strftime("%Y%m%dT%H%M%S")),
                "DTEND;TZID=%s:%s" % (TIMEZONE, ev["end"].strftime("%Y%m%dT%H%M%S")),
            ]
        else:
            nxt = ev["date"] + dt.timedelta(days=1)
            lines += [
                "DTSTART;VALUE=DATE:" + ev["date"].strftime("%Y%m%d"),
                "DTEND;VALUE=DATE:" + nxt.strftime("%Y%m%d"),
            ]
        lines.append(fold("SUMMARY:" + esc(ev["summary"])))
        if ev["note"]:
            lines.append(fold("DESCRIPTION:" + esc("Roster note: " + ev["note"])))
        lines += ["TRANSP:OPAQUE", "END:VEVENT"]
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--person", default="Elaine")
    ap.add_argument("--out", default="roster.ics")
    ap.add_argument("--years", default="", help="comma-separated year tabs (default: all year tabs)")
    ap.add_argument("--no-leave", action="store_true", help="shifts only")
    ap.add_argument("--from", dest="date_from", default="2026-01-01", metavar="YYYY-MM-DD",
                    help="ignore days before this (default: 2026-01-01)")
    ap.add_argument("--to", dest="date_to", default="2027-03-31", metavar="YYYY-MM-DD",
                    help="ignore days after this, i.e. where the roster stops being "
                         "accurate (default: 2027-03-31)")
    args = ap.parse_args()

    def bound(text, name):
        if not text:
            return None
        try:
            return dt.datetime.strptime(text, "%Y-%m-%d").date()
        except ValueError:
            sys.exit("--%s must look like YYYY-MM-DD" % name)

    date_from = bound(args.date_from, "from")
    date_to = bound(args.date_to, "to")

    wb = openpyxl.load_workbook(args.workbook)
    wanted = [y.strip() for y in args.years.split(",") if y.strip()]
    tabs = [t for t in wb.sheetnames if re.fullmatch(r"\d{4}", t) and (not wanted or t in wanted)]
    if not tabs:
        sys.exit("no year tabs found")

    records = []
    for tab in tabs:
        records += list(collect(wb[tab], int(tab), args.person))
    if not records:
        sys.exit("no rows found for %s in tabs %s" % (args.person, ", ".join(tabs)))

    kept = [r for r in records
            if (date_from is None or r[0] >= date_from)
            and (date_to is None or r[0] <= date_to)]
    dropped = len(records) - len(kept)
    records = kept

    events = build_events(records, include_leave=not args.no_leave)
    events.sort(key=lambda e: (e["date"], e["kind"]))
    stamp = dt.datetime.now(dt.timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    with open(args.out, "w", newline="") as fh:
        fh.write(to_ics(events, args.person, stamp))

    timed = sum(1 for e in events if e["kind"] == "timed")
    print("tabs: %s" % ", ".join(tabs))
    print("window: %s to %s (%d days outside it ignored)"
          % (date_from or "start", date_to or "end", dropped))
    print("%d events (%d shifts, %d all-day) -> %s" % (len(events), timed, len(events) - timed, args.out))


if __name__ == "__main__":
    main()
