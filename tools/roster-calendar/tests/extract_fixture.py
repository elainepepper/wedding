#!/usr/bin/env python3
"""Dump the roster workbook into fixture.json for run_tests.js.

Captures what the Sheets API would hand the Apps Script - values, display
values and background colours - so the script can be exercised offline.

    python3 extract_fixture.py Tech_Roster.xlsx [--years 2026,2027]

The output contains real roster data, so it is gitignored.
"""

import argparse
import datetime
import json
import os
import re

import openpyxl

COLUMNS = 32  # column A plus 31 day columns, matching DAY_COLUMNS in the script
MAX_ROWS = 400


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("workbook")
    ap.add_argument("--years", default="", help="comma-separated year tabs (default: all)")
    ap.add_argument("--out", default=os.path.join(os.path.dirname(__file__), "fixture.json"))
    args = ap.parse_args()

    wanted = [y.strip() for y in args.years.split(",") if y.strip()]
    wb = openpyxl.load_workbook(args.workbook)
    out = {}

    for name in wb.sheetnames:
        if not re.fullmatch(r"\d{4}", name):
            continue
        if wanted and name not in wanted:
            continue
        ws = wb[name]
        values, display, backgrounds = [], [], []
        for r in range(1, min(ws.max_row, MAX_ROWS) + 1):
            vrow, drow, brow = [], [], []
            for c in range(1, COLUMNS + 1):
                cell = ws.cell(r, c)
                v = cell.value
                if isinstance(v, (datetime.datetime, datetime.date)):
                    vrow.append({"__date": v.isoformat()})
                    drow.append(v.strftime("%d/%m/%Y"))
                elif v is None:
                    vrow.append("")
                    drow.append("")
                elif isinstance(v, bool):
                    vrow.append(v)
                    drow.append(str(v))
                elif isinstance(v, (int, float)):
                    vrow.append(v)
                    drow.append("%g" % v)
                else:
                    vrow.append(str(v))
                    drow.append(str(v))

                fill = cell.fill
                rgb = fill.start_color.rgb if (fill and fill.patternType) else None
                if not isinstance(rgb, str):
                    brow.append("#ffffff")
                else:
                    rgb = rgb.lower()
                    brow.append("#" + (rgb[2:] if len(rgb) == 8 else rgb))
            values.append(vrow)
            display.append(drow)
            backgrounds.append(brow)
        out[name] = {"values": values, "display": display, "backgrounds": backgrounds}

    if not out:
        raise SystemExit("no year tabs found")
    with open(args.out, "w") as fh:
        json.dump(out, fh)
    print("tabs: %s -> %s" % (", ".join(out), args.out))


if __name__ == "__main__":
    main()
